import openpyxl, pickle, re
from pathlib import Path
import os, anytree, logging
# import pyinputplus as pyip
# from lib.creoClass import CreoNode


BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__))) # Get working directory of file
os.chdir(BASE_DIR)

TREE_BINARY = Path('./Excel/7u-bottom.bom.1.pk')
IMAGE_DIR = Path('./7U-Bottom/images_1000')

LOGPATH = Path('./logs/')
logname = Path('bom_excel_export.txt')

# try:
#     os.remove(logpath + logname) # clears previous log file
#     print('Clearing Logfile')
# except:
#     print(f'Creating Logfile: {logpath}{logname}')
logging.basicConfig(filename=LOGPATH / logname, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable()

def importPickleBinary():
    '''Opens binary file of path TREE_BINARY and returns CreoNode object'''
    with open(TREE_BINARY, 'rb') as input:
        creoNode = pickle.load(input)
    return creoNode

def searchFileImg(partName, partType, directory):
    '''Searches "directory" for image file using part name and type'''

    imageNameFiles = os.listdir(directory)
    
    regexImage = re.compile((f'{str(partName)}') + r'(prt|asm).(jpg|png|jpeg|)', re.IGNORECASE)
    for file in imageNameFiles:
        result = regexImage.match(file)
        if result:
            filePartType = result[1]
            if filePartType.lower() == partType.lower():
                logging.info(f'Found match for {partName}.{partType} = {result[0]}')
                return result[0]
    logging.critical(f'Found no match for {partName}.{partType}')
    return None

if __name__ == '__main__':

    asmNode = importPickleBinary()

    # Declare image size variables and rough fudge factor to excel widths
    IMAGE_WIDTH = None
    IMAGE_HEIGHT = None
    IMAGE_SCALE = .12 # Approximate scale for 1000 x 1000 pixel image, change this to scale images in document
    WIDTH_FUDGE = (1/7)
    HEIGHT_FUDGE = (3/4)

    # Expand tree in order from top to bottom as line items
    treeDepth = asmNode.height
    wrkb = openpyxl.Workbook()     # # Create workbook class
    ws = wrkb.worksheets[0] # Number of sheets in the workbook (1 sheet in our case)

    # Print Rows: Name / Image / QTY / FULLQTY / LEVELS(1-N:(max depth))
    asmColumns = (treeDepth + 1) * ['']
    rowCount = 0

    # Print Header to worksheet
    headerCol = asmColumns[:]
    for i in range(len(headerCol)): headerCol[i] = f'LVL: {i + 1}'
    header = ['Name', 'Image', 'QTY', 'BRANCH\nQTY','TOTAL\nQTY', 'Path', 'Depth'] + headerCol
    ws.append(header)

    # Print data to worksheet
    rowIndex = 1
    for node in anytree.PreOrderIter(asmNode):
        
        # Print Data
        col_name = node.name + '.' + node.type
        col_image = ''
        col_qty = node.qty
        col_brch_qty = node.branchQTY
        col_fullQty = node.totalTreeQTY
        col_path = node.getParentsPrintout()
        rowAsmColumns = asmColumns[:]
        rowAsmColumns[node.depth] = 'X'
        row = [str(col_name), col_image, int(col_qty), int(col_brch_qty), int(col_fullQty), col_path, int(node.depth+1)] + rowAsmColumns
        ws.append(row)

        # Add file image
        imgName = searchFileImg(node.name, node.type, IMAGE_DIR)
        if imgName != None:
            img = openpyxl.drawing.image.Image(IMAGE_DIR / imgName)
            img.anchor = 'B' + str(rowIndex+1)
            img.height = img.height*IMAGE_SCALE # insert image height in pixels as float or int (e.g. 305.5)
            img.width= img.width*IMAGE_SCALE # insert image width in pixels as float or int (e.g. 405.8)
            IMAGE_WIDTH = img.height # In pixels
            IMAGE_HEIGHT = img.width # In pixels
            ws.add_image(img)
        else:
            print(f'No image found for {node.name}.{node.type}')

        rowIndex += 1

    # Resize columns and rows for images

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = IMAGE_WIDTH * WIDTH_FUDGE + 1

    for row in range(ws.max_row+1):
        if row == 0:
            ws.row_dimensions[row].height = 20
        if row > 1:
            ws.row_dimensions[row].height = IMAGE_HEIGHT * HEIGHT_FUDGE + 1

    # Misc formatting
    cell = ws['B2'] ; ws.freeze_panes = cell # Freeze top row
    ws.auto_filter.ref = ws.dimensions  # Add filter to data

    wrkb.save('excel_bom_output.xlsx')