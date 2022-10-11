import openpyxl, pickle, re
from pathlib import Path
import os, anytree, logging

# Declare basepath and filenames
BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__))) # Get working directory of file
os.chdir(BASE_DIR)

# Filename declarations
LOG_FILE = Path('bom_frontend.py.txt')
IMAGE_PATH = Path('./creo_images')
BINARY_IMP_FILE = Path('BINARY_EXPORT.pk')

# Get current working directory
BASE_DIR = Path(os.path.dirname(os.path.abspath(__file__))) # Get working directory of file
os.chdir(BASE_DIR)

# Folder locations for data
LOG_ROOT_PATH = Path('./logs')
EXPORT_ROOT_PATH = Path('./exports')
IMAGES_PATH = Path('./creo_images')

# Export locations
LOG_PATH = LOG_ROOT_PATH / LOG_FILE
EXCEL_EXP_PATH = './exports/EXCEL_EXPORT.xlsx'
BINARY_IMP_PATH = EXPORT_ROOT_PATH / BINARY_IMP_FILE

# Logging variables
try:
    os.remove(LOG_PATH) # clears previous log file
except:
    print(f'Creating Logfile: {LOG_FILE}')
logging.basicConfig(filename=LOG_PATH, level=logging.CRITICAL, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable()

def importPickleBinary():
    '''Opens binary file of path TREE_BINARY and returns CreoNode object'''
    with BINARY_IMP_PATH.open('rb') as input:
        creoNode, asmBOM = pickle.load(input)
    return creoNode, asmBOM

def searchFileImg(partName, partType, directory):
    '''Searches "directory" for image file using part name and type'''

    imageNameFiles = os.listdir(directory)
    
    regexImage = re.compile((f'{str(partName)}') + r'(prt|asm).(jpg|png|jpeg|)', re.IGNORECASE)
    for file in imageNameFiles:
        result = regexImage.match(file)
        if result:
            filePartType = result[1]
            if filePartType.lower() == partType.lower():
                logging.info(f'Found match for {partName}.{partType} = {result[0]}')
                return result[0]
    logging.critical(f'Found no match for {partName}.{partType}')
    return None

if __name__ == '__main__':

    asmNode, asmBOM = importPickleBinary()

    # Declare image size variables and rough fudge factor to excel widths
    IMAGE_WIDTH = None
    IMAGE_HEIGHT = None
    IMAGE_SCALE = .12 # Approximate scale for 1000 x 1000 pixel image, change this to scale images in document
    WIDTH_FUDGE = (1/7)
    HEIGHT_FUDGE = (3/4)

    # Expand tree in order from top to bottom as line items
    treeDepth = asmNode.height
    wrkb = openpyxl.Workbook()     # # Create workbook class
    ws1 = wrkb.worksheets[0] # Number of sheets in the workbook (1 sheet in our case)
    ws1.title = 'Full_BOM'

    # Print Rows: Name / Image / QTY / FULLQTY / LEVELS(1-N:(max depth))
    asmColumns = (treeDepth + 1) * ['']
    rowCount = 0

    # Print Header to worksheet
    headerCol = asmColumns[:]
    for i in range(len(headerCol)): headerCol[i] = f'LVL: {i + 1}'
    header = ['Name', 'Image', 'QTY', 'BRANCH\nQTY','TOTAL\nQTY', 'Path', 'Depth'] + headerCol
    ws1.append(header)

    # Print data to worksheet
    rowIndex = 1
    for node in anytree.PreOrderIter(asmNode):
        
        # Print Data
        col_name = node.name + '.' + node.type
        col_image = ''
        col_qty = node.qty
        col_brch_qty = node.branchQTY
        col_fullQty = node.totalTreeQTY
        col_path = node.getParentsPrintout()
        rowAsmColumns = asmColumns[:]
        rowAsmColumns[node.depth] = 'X'
        row = [str(col_name), col_image, int(col_qty), int(col_brch_qty), int(col_fullQty), col_path, int(node.depth+1)] + rowAsmColumns
        ws1.append(row)

        # Add file image
        imgName = searchFileImg(node.name, node.type, IMAGES_PATH)
        if imgName != None:
            img = openpyxl.drawing.image.Image(IMAGES_PATH / imgName)
            img.anchor = 'B' + str(rowIndex+1)
            img.height = img.height*IMAGE_SCALE # insert image height in pixels as float or int (e.g. 305.5)
            img.width= img.width*IMAGE_SCALE # insert image width in pixels as float or int (e.g. 405.8)
            IMAGE_WIDTH = img.height # In pixels
            IMAGE_HEIGHT = img.width # In pixels
            ws1.add_image(img)
        else:
            print(f'No image found for {node.name}.{node.type}')

        rowIndex += 1

    # Resize columns and rows for images
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = IMAGE_WIDTH * WIDTH_FUDGE + 1

    for row in range(ws1.max_row+1):
        if row == 0:
            ws1.row_dimensions[row].height = 20
        if row > 1:
            ws1.row_dimensions[row].height = IMAGE_HEIGHT * HEIGHT_FUDGE + 1

    # Misc formatting
    cell = ws1['B2'] ; ws1.freeze_panes = cell # Freeze top row
    ws1.auto_filter.ref = ws1.dimensions  # Add filter to data

    ws2 = wrkb.create_sheet()
    ws2.title = 'Consolidated_BOM'

    ws2Header = ['Name', 'Type', 'Qty']
    ws2.append(ws2Header)

    for creoPart in asmBOM:
        row = [creoPart[2], creoPart[1], int(creoPart[0])]
        ws2.append(row)

    wrkb.save(EXCEL_EXP_PATH)