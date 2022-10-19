import logging
import os
import re
from datetime import datetime
from pathlib import Path

import anytree
import openpyxl
from anytree.exporter import JsonExporter

from . import constants
from .creoClass import CreoNode


def searchFileImg(partName, partType, directory):
    """Searches "directory" for image file using part name and type"""

    imageNameFiles = os.listdir(directory)

    regexImage = re.compile(
        (f"{str(partName)}") + r"(prt|asm).(jpg|png|jpeg|)", re.IGNORECASE
    )
    # for file in imageNameFiles:
    for root, dirs, files in os.walk(constants.IMAGES_DIR):
        for name in files:
            result = regexImage.match(name)
            if result:
                filePartType = result[1]
                if filePartType.lower() == partType.lower():
                    logging.info(f"Found match for {partName}.{partType} = {result[0]}")
                    return os.path.join(root, name)
    logging.warning(f"Found no match for {partName}.{partType}")


def exportExcel(asmNode, asmBOM, filenameprefix, filenamesuffix):

    # Declare image size variables and rough fudge factor to excel widths
    IMAGE_WIDTH = None
    IMAGE_HEIGHT = None
    IMAGE_SCALE = 0.12  # Approximate scale for 1000 x 1000 pixel image, change this to scale images in document
    WIDTH_FUDGE = 1 / 7
    HEIGHT_FUDGE = 3 / 4

    # Expand tree in order from top to bottom as line items
    treeDepth = asmNode.height
    wrkb = openpyxl.Workbook()  # # Create workbook class
    ws1 = wrkb.worksheets[0]  # Number of sheets in the workbook (1 sheet in our case)
    now = datetime.now()
    ws1.title = "asmNode.name" + "_" + now.strftime("%y%d%m_%H%M")

    # Print Rows: Name / Image / QTY / FULLQTY / LEVELS(1-N:(max depth))
    asmColumns = (treeDepth + 1) * [""]
    rowCount = 0

    # Print Header to worksheet
    headerCol = asmColumns[:]
    for i in range(len(headerCol)):
        headerCol[i] = f"LVL: {i + 1}"
    header = [
        "Name",
        "Image",
        "QTY",
        "BRANCH\nQTY",
        "TOTAL\nQTY",
        "Path",
        "Depth",
    ] + headerCol
    ws1.append(header)  # type: ignore

    # Print data to worksheet
    rowIndex = 1
    for node in anytree.PreOrderIter(asmNode):

        # Print Data
        col_name = node.name + "." + node.type
        col_image = ""
        col_qty = node.qty
        col_brch_qty = node.branchQTY
        col_fullQty = node.totalTreeQTY
        col_path = node.getParentsPrintout()
        rowAsmColumns = asmColumns[:]
        rowAsmColumns[node.depth] = "X"
        row = [
            str(col_name),
            col_image,
            int(col_qty),
            int(col_brch_qty),
            int(col_fullQty),
            col_path,
            int(node.depth + 1),
        ] + rowAsmColumns
        ws1.append(row)  # type: ignore

        # Add file image
        imgPath = searchFileImg(node.name, node.type, str(constants.IMAGES_DIR))
        if imgPath != None:
            img = openpyxl.drawing.image.Image(imgPath)  # type: ignore
            img.anchor = "B" + str(rowIndex + 1)
            img.height = (
                img.height * IMAGE_SCALE
            )  # insert image height in pixels as float or int (e.g. 305.5)
            img.width = (
                img.width * IMAGE_SCALE
            )  # insert image width in pixels as float or int (e.g. 405.8)
            IMAGE_WIDTH = img.height  # In pixels
            IMAGE_HEIGHT = img.width  # In pixels
            ws1.add_image(img)  # type: ignore
        else:
            print(f"No image found for {node.name}.{node.type}")

        rowIndex += 1

    # Resize columns and rows for images
    ws1.column_dimensions["A"].width = 40  # type: ignore
    ws1.column_dimensions["B"].width = IMAGE_WIDTH * WIDTH_FUDGE + 1  # type: ignore

    for row in range(ws1.max_row + 1):
        if row == 0:
            ws1.row_dimensions[row].height = 20  # type: ignore
        if row > 1:
            ws1.row_dimensions[row].height = IMAGE_HEIGHT * HEIGHT_FUDGE + 1  # type: ignore

    # Misc formatting
    cell = ws1["B2"]
    ws1.freeze_panes = cell  # type: ignore # Freeze top row
    ws1.auto_filter.ref = ws1.dimensions  # type: ignore # Add filter to data

    ws2 = wrkb.create_sheet()
    ws2.title = "Consolidated_BOM"

    ws2Header = ["Name", "Type", "Qty"]
    ws2.append(ws2Header)

    for creoPart in asmBOM:
        row = [creoPart[2], creoPart[1], int(creoPart[0])]
        ws2.append(row)

    XLS_EXP_PATH = constants.XLS_EXP_DIR / Path(filenameprefix + filenamesuffix + ".XLSX")

    logging.info(f"Writing Excel BOM to file {str(XLS_EXP_PATH)}")
    print(f"Writing Excel BOM to file {str(XLS_EXP_PATH)}")
    wrkb.save(str(XLS_EXP_PATH))


def exportJSON(activeNode, filenameprefix, filenamesuffix):
    # Data and file export to JSON
    exporter = JsonExporter(indent=2, sort_keys=True)
    JSON_EXP_PATH = constants.JSON_EXP_DIR / Path(filenameprefix + filenamesuffix + ".JSON")
    logging.info(f"Writing JSON to file {str(JSON_EXP_PATH)}")
    print(f"Writing JSON to file {str(JSON_EXP_PATH)}")
    with JSON_EXP_PATH.open("w", encoding="utf-8-sig") as fileObject:
        fileObject.write(exporter.export(activeNode))


if __name__ == "__main__":
    print("Not indended to run standalone")
